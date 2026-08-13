from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def export_attendance(records, report_date):
    """
    Exports attendance records to an Excel file.
    """

    if len(records) == 0:
        return False, "No attendance records found."

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Attendance Report"

    # -------------------------
    # Report Title
    # -------------------------

    sheet["A1"] = f"Attendance Report - {report_date}"
    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    # -------------------------
    # Headers
    # -------------------------

    headers = [
        "Roll Number",
        "Student Name",
        "Attendance Time"
    ]

    for column, header in enumerate(headers, start=1):

        cell = sheet.cell(
            row=3,
            column=column
        )

        cell.value = header
        cell.font = Font(bold=True)

    # -------------------------
    # Attendance Data
    # -------------------------

    for row, record in enumerate(records, start=4):

        sheet.cell(row=row, column=1).value = record[0]
        sheet.cell(row=row, column=2).value = record[1]
        sheet.cell(row=row, column=3).value = record[2]

    # -------------------------
    # Auto Width
    # -------------------------

    for column_cells in sheet.columns:

        max_length = 0

        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        sheet.column_dimensions[column_letter].width = max_length + 5

    # -------------------------
    # Create Reports Folder
    # -------------------------

    reports_folder = Path("reports")
    reports_folder.mkdir(exist_ok=True)

    filename = reports_folder / f"Attendance_{report_date}.xlsx"

    workbook.save(filename)

    return True, filename